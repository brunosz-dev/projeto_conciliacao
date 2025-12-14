"""
Testes para o módulo excel_writer.py
"""
import pytest
import openpyxl
from src.excel_writer import escrever_resultados

MOCK_DADOS = [
    {
        'id_venda': 'TX-99',
        'cliente': 'Fulano',
        'valor_bruto': 100.0,
        'forma_pagamento': 'pix',
        'taxa_gateway': 1.0,
        'taxa_adicional': 0.5,
        'valor_liquido': 98.5,
        'custo_produto': 50.0,
        'lucro': 48.5,
        'roi': 97.0,
        'status': 'Aprovado'
    }
]

@pytest.mark.io
def test_escrever_resultados_gera_arquivo(tmp_path):
    """Verifica se o arquivo é criado e o conteúdo está correto."""
    # 1. Arrange
    arquivo_saida = tmp_path / "relatorio_teste.xlsx"
    
    # 2. Act
    escrever_resultados(MOCK_DADOS, str(arquivo_saida))
    
    # 3. Assert
    assert arquivo_saida.exists()
    
    # Reabre para conferir conteúdo
    wb = openpyxl.load_workbook(arquivo_saida)
    ws = wb.active
    
    # Cabeçalho está na linha 1?
    assert ws['A1'].value == "ID Venda"
    
    # Dados estão na linha 2?
    assert ws['A2'].value == "TX-99"
    assert ws['C2'].value == 100.0  # Valor Bruto
    assert ws['K2'].value == "Aprovado" # Status
    
    # Verifica linha de Totais (Linha 3 neste caso)
    assert ws['A3'].value == "TOTAL"
    assert ws['C3'].value == 100.0 # Soma do bruto

@pytest.mark.io
def test_escrever_resultados_formatacao(tmp_path):
    """Verifica (basicamente) se formatação condicional foi aplicada."""
    arquivo_saida = tmp_path / "relatorio_cores.xlsx"
    escrever_resultados(MOCK_DADOS, str(arquivo_saida))
    
    wb = openpyxl.load_workbook(arquivo_saida)
    ws = wb.active
    
    cell_status = ws['K2']
    # O excel_writer usa PatternFill. Vamos checar se o fillType é solid
    # (Teste profundo de cor exata hexadecimal é frágil, testamos a aplicação do estilo)
    assert cell_status.fill.patternType == 'solid'