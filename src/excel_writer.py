"""
Módulo responsável pela escrita do relatório final em Excel.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)


# Cores profissionais
COLORS = {
    "header": "4472C4",      # Azul
    "aprovado": "C6EFCE",    # Verde claro
    "pendente": "FFEB9C",    # Amarelo
    "cancelado": "FFC7CE",   # Vermelho claro
    "total": "FFC000"        # Laranja
}


def escrever_resultados(dados: list, output_file: str) -> None:
    """
    Escreve os resultados processados em um arquivo Excel formatado.
    
    Args:
        dados: Lista de dicionários com os resultados
        output_file: Caminho do arquivo de saída
        
    Example:
        >>> dados = [{'cliente': 'João', 'valor_bruto': 100, ...}]
        >>> escrever_resultados(dados, 'output/relatorio.xlsx')
    """
    try:
        logger.info(f"📝 Criando relatório: {output_file}")
        
        # Criar workbook
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Relatório de Conciliação"
        
        # Escrever cabeçalho
        headers = [
            "ID Venda",
            "Cliente",
            "Valor Bruto",
            "Forma Pagamento",
            "Taxa Gateway",
            "Taxa Adicional",
            "Valor Líquido",
            "Custo",
            "Lucro",
            "ROI (%)",
            "Status"
        ]
        
        _escrever_cabecalho(sheet, headers)
        
        # Escrever dados
        for row_idx, item in enumerate(dados, start=2):
            _escrever_linha(sheet, row_idx, item)
        
        # Adicionar linha de totais
        _adicionar_totais(sheet, len(dados) + 2, dados)
        
        # Ajustar larguras
        _ajustar_larguras(sheet)
        
        # Salvar
        wb.save(output_file)
        logger.info(f"✅ Relatório salvo com sucesso!")
        
    except Exception as e:
        logger.error(f"❌ Erro ao escrever Excel: {str(e)}")
        raise


def _escrever_cabecalho(sheet, headers: list) -> None:
    """Escreve e formata o cabeçalho."""
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=COLORS["header"], 
                                end_color=COLORS["header"], 
                                fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )


def _escrever_linha(sheet, row_idx: int, item: dict) -> None:
    """Escreve uma linha de dados com formatação."""
    valores = [
        item.get('id_venda', ''),
        item.get('cliente', ''),
        item.get('valor_bruto', 0),
        item.get('forma_pagamento', ''),
        item.get('taxa_gateway', 0),
        item.get('taxa_adicional', 0),
        item.get('valor_liquido', 0),
        item.get('custo_produto', 0),
        item.get('lucro', 0),
        item.get('roi', 0),
        item.get('status', 'Pendente')
    ]
    
    for col_idx, valor in enumerate(valores, start=1):
        cell = sheet.cell(row=row_idx, column=col_idx)
        cell.value = valor
        
        # Formatação de moeda
        if col_idx in [3, 5, 6, 7, 8, 9]:  # Colunas de valor
            cell.number_format = 'R$ #,##0.00'
        
        # Formatação de percentual
        if col_idx == 10:  # ROI
            cell.number_format = '0.00"%"'
        
        # Cor baseada no status
        if col_idx == 11:  # Status
            status = str(valor).lower()
            if "aprovado" in status:
                cell.fill = PatternFill(start_color=COLORS["aprovado"], 
                                       end_color=COLORS["aprovado"], 
                                       fill_type="solid")
            elif "pendente" in status:
                cell.fill = PatternFill(start_color=COLORS["pendente"], 
                                       end_color=COLORS["pendente"], 
                                       fill_type="solid")


def _adicionar_totais(sheet, row_idx: int, dados: list) -> None:
    """Adiciona linha de totais."""
    sheet.cell(row=row_idx, column=1).value = "TOTAL"
    sheet.cell(row=row_idx, column=1).font = Font(bold=True)
    
    # Somar valores
    total_bruto = sum(item.get('valor_bruto', 0) for item in dados)
    total_liquido = sum(item.get('valor_liquido', 0) for item in dados)
    total_lucro = sum(item.get('lucro', 0) for item in dados)
    
    sheet.cell(row=row_idx, column=3).value = total_bruto
    sheet.cell(row=row_idx, column=7).value = total_liquido
    sheet.cell(row=row_idx, column=9).value = total_lucro
    
    # Formatação da linha de total
    for col in range(1, 12):
        cell = sheet.cell(row=row_idx, column=col)
        cell.fill = PatternFill(start_color=COLORS["total"], 
                               end_color=COLORS["total"], 
                               fill_type="solid")
        cell.font = Font(bold=True)


def _ajustar_larguras(sheet) -> None:
    """Ajusta automaticamente as larguras das colunas."""
    for column_cells in sheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2